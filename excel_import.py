"""📥 Importar Excel — bulk load of Anexo 2 production workbooks (R5).

A 4-step wizard (subir → mapear catálogos → previsualizar → confirmar) that loads
Masivos/Bitácoras rows into faena + children. Orchestration + UI only; parsing lives
in import_formats, catalog resolution in catalog_resolver, insertion in import_writer.
"""
from __future__ import annotations
import streamlit as st

import import_formats as IF
import catalog_resolver as R

STEPS = ["Subir", "Mapear catálogos", "Previsualizar", "Confirmar"]


def _step() -> int:
    return st.session_state.setdefault("imp_step", 0)


def _reset():
    for k in list(st.session_state):
        if k.startswith("imp_"):
            del st.session_state[k]


def render_excel_import():
    from console_ui import page_header
    page_header(
        "📥 Importar Excel",
        "Carga masiva de datos históricos (formatos Masivos y Bitácoras del Anexo 2).",
        help_md=(
            "1. **Sube** el archivo `.xlsx` y confirma el formato detectado.\n"
            "2. **Mapea** los nombres del archivo a los catálogos (acepta o corrige las sugerencias).\n"
            "3. **Previsualiza**: cuántas faenas nuevas, cuáles ya existen, errores.\n"
            "4. **Confirma** para guardar. Los catálogos nuevos quedan sin aprobar, para revisión."
        ),
    )
    st.progress((_step()) / (len(STEPS) - 1), text=f"Paso {_step()+1}/{len(STEPS)}: {STEPS[_step()]}")
    if st.button("↺ Empezar de nuevo", key="imp_restart"):
        _reset(); st.rerun()

    if _step() == 0:
        _step1_upload()
    elif _step() == 1:
        _step2_map()
    elif _step() == 2:
        _step3_preview()
    else:
        _step4_commit()


def _distinct_catalog_values(rows, spec):
    vals = {}      # (catalog, raw) -> None
    for r in rows:
        for header, t in {**spec.faena_cols, **spec.catch_cols}.items():
            if t.kind == "catalog" and not R.is_na(r.get(header)):
                vals[(t.catalog, R.normalize(r.get(header)))] = None
        # child catalogs (arte, carnada sitio/arte)
        for col, (header, cat) in {**spec.children["arte"]}.items():
            if cat and not R.is_na(r.get(header)):
                vals[(cat, R.normalize(r.get(header)))] = None
    return list(vals)


def build_mapping_model(rows, spec):
    values = [(c, v) for (c, v) in _distinct_catalog_values(rows, spec)]
    especies = []
    seen = set()
    for r in rows:
        pair = (R.normalize(r.get(spec.especie_comun)), R.normalize(r.get(spec.especie_cientifico)))
        if pair != ("", "") and pair not in seen:
            seen.add(pair); especies.append(pair)
    return {"values": {kv: None for kv in values}, "especies": especies}


def _step1_upload():
    import openpyxl
    up = st.file_uploader("Archivo Excel (.xlsx)", type=["xlsx"], key="imp_file")
    if not up:
        return
    wb = openpyxl.load_workbook(up, read_only=True, data_only=True)
    # Read every sheet's header row and auto-detect its format. The workbook has several sheets
    # with near-identical headers (Masivos/Bitácoras) plus non-data sheets, so the admin picks
    # which sheet to import; detection only pre-selects the sensible default.
    sheets = {}
    for ws in wb.worksheets:
        headers = [c for c in next(ws.iter_rows(values_only=True), [])]
        sheets[ws.title] = (headers, IF.detect_format([h for h in headers if h]))
    names = list(sheets)
    default_idx = next((i for i, n in enumerate(names) if sheets[n][1]), 0)
    title = st.selectbox("Hoja del archivo", names, index=default_idx)
    headers, code_guess = sheets[title]
    if not any(h for h in headers):
        st.error("Esa hoja no tiene encabezados en la primera fila."); return
    fmt_names = list(IF.FORMATS)
    code = st.selectbox("Formato", fmt_names,
                        index=fmt_names.index(code_guess) if code_guess else 0,
                        format_func=lambda c: IF.FORMATS[c].codigo)
    if code_guess is None:
        st.warning("No detecté el formato de esta hoja automáticamente — confírmalo arriba.")
    ws = wb[title]
    data = list(ws.iter_rows(min_row=2, values_only=True))
    rows = IF.parse_rows(headers, data)
    st.success(f"Hoja **{title}** · formato **{code}** · **{len(rows)}** filas de datos.")
    if st.button("Continuar →", type="primary"):
        st.session_state["imp_format"] = code
        st.session_state["imp_rows"] = rows
        st.session_state["imp_step"] = 1
        st.rerun()


def _step2_map():
    spec = IF.FORMATS[st.session_state["imp_format"]]
    rows = st.session_state["imp_rows"]
    model = build_mapping_model(rows, spec)
    st.caption("Confirma a qué catálogo corresponde cada nombre. Las coincidencias exactas ya están "
               "resueltas; revisa sólo lo que no coincide.")
    mapping = st.session_state.setdefault("imp_map", {})
    unresolved = 0
    for (catalog, raw), _ in model["values"].items():
        exact = R.resolve_exact(catalog, raw)
        if exact:
            mapping[(catalog, raw)] = exact; continue
        unresolved += 1
        with st.container(border=True):
            st.markdown(f"**{raw}** · `{catalog}`")
            sugg = R.fuzzy_suggest(catalog, raw)
            choices = {f"➕ Crear «{raw}»": ("new", raw), "🚫 Desconocido": ("desc", None)}
            for cid, name, score in sugg:
                choices[f"{name}  ({int(score*100)}%)"] = ("id", cid)
            pick = st.selectbox("Asignar a", list(choices), key=f"imp_m_{catalog}_{raw}")
            kind, val = choices[pick]
            if kind == "new":
                mapping[(catalog, raw)] = ("__NEW__", raw)
            elif kind == "desc":
                mapping[(catalog, raw)] = ("__DESC__", catalog)
            else:
                mapping[(catalog, raw)] = val
    st.caption(f"{unresolved} valor(es) por confirmar. Especies se resuelven por par común+científico "
               "en la vista previa.")
    c1, c2 = st.columns(2)
    if c1.button("← Volver"):
        st.session_state["imp_step"] = 0; st.rerun()
    if c2.button("Previsualizar →", type="primary"):
        st.session_state["imp_step"] = 2; st.rerun()


import import_writer as IW


def apply_mapping(spec, drafts, mapping):
    """Rewrite each draft's ('catalog', cat, raw) cells using the admin mapping, then resolve
    the rest (create/Desconocido/especie) via import_writer.resolve_draft."""
    def rewrite(v):
        if isinstance(v, tuple) and len(v) == 3 and v[0] == "catalog":
            _, cat, raw = v
            choice = mapping.get((cat, R.normalize(raw)))
            if isinstance(choice, str):                      # a chosen existing id
                return choice
            if isinstance(choice, tuple) and choice[0] == "__DESC__":
                return R.desconocido_id(cat)
            return R.resolve_or_create(cat, raw)             # __NEW__ or unmapped → create/exact
        return v
    resolved = []
    for d in drafts:
        d.faena_raw = {k: rewrite(v) for k, v in d.faena_raw.items()}
        d.children_raw["arte"] = {k: rewrite(v) for k, v in d.children_raw["arte"].items()}
        resolved.append(IW.resolve_draft(spec, d))
    return resolved


def _step3_preview():
    spec = IF.FORMATS[st.session_state["imp_format"]]
    drafts = IF.group_faenas(st.session_state["imp_rows"], spec)
    resolved = apply_mapping(spec, drafts, st.session_state.get("imp_map", {}))
    hashes = [r["faena"]["legacy_id"] for r in resolved if r["key"]]
    dup = IW.existing_legacy_ids(hashes)
    nuevas = sum(1 for r in resolved if r["key"] and r["faena"]["legacy_id"] not in dup)
    ya = sum(1 for r in resolved if r["key"] and r["faena"]["legacy_id"] in dup)
    err = sum(1 for r in resolved if r["key"] is None)
    nueva = lambda r: r["key"] and r["faena"]["legacy_id"] not in dup
    if spec.kind == "monitoreo":
        hijos = sum(len(r.get("mediciones", [])) for r in resolved if nueva(r)); hijos_label = "Mediciones"
    else:
        hijos = sum(len(r["catches"]) for r in resolved if nueva(r)); hijos_label = "Capturas"
    st.session_state["imp_resolved"] = resolved
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Faenas nuevas", nuevas); c2.metric("Ya existen", ya)
    c3.metric(hijos_label, hijos); c4.metric("Con error", err)
    warnings = [e for r in resolved for e in r["errors"]]
    if warnings:
        with st.expander(f"⚠️ {len(warnings)} avisos"):
            for w in warnings[:200]:
                st.caption("• " + w)
    force = st.checkbox("Forzar inclusión de faenas que ya existen", key="imp_force")
    c1, c2 = st.columns(2)
    if c1.button("← Volver"):
        st.session_state["imp_step"] = 1; st.rerun()
    if c2.button(f"Guardar {nuevas} faena(s) →", type="primary", disabled=nuevas == 0 and not force):
        st.session_state["imp_step"] = 3; st.rerun()


def _step4_commit():
    from console_ui import friendly_error
    spec = IF.FORMATS[st.session_state["imp_format"]]
    resolved = st.session_state["imp_resolved"]
    try:
        rep = IW.commit_batch(spec, resolved, force=st.session_state.get("imp_force", False))
    except Exception as e:                                    # noqa: BLE001
        st.error(friendly_error(e)); return
    hijos = (f"{rep['mediciones']} mediciones" if spec.kind == "monitoreo"
             else f"{rep['capturas']} capturas")
    st.success(f"✅ {rep['faenas_nuevas']} faenas · {hijos} guardadas. "
               f"{rep['ya_existen']} ya existían. {rep['faenas_error']} con error.")
    if rep["errores"]:
        with st.expander("Errores"):
            for e in rep["errores"][:200]:
                st.caption("• " + e)
    st.info("Los catálogos nuevos quedaron **sin aprobar** — revísalos en 🔎 Duplicados / 📥 Propuestas.")
    if st.button("Importar otro archivo", type="primary"):
        _reset(); st.rerun()
